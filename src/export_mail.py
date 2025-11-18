"""
export_mail.py - Lógica de exportación y envío de reportes por mail
"""

import smtplib
import json
import os
import sys
from datetime import datetime
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email.mime.text import MIMEText
from email import encoders
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

from config import UPLOAD_FOLDER, MAIL_SERVER, MAIL_PORT, MAIL_USERNAME, MAIL_PASSWORD
from src.db import get_db


def export_monthly(email):
    """
    Genera reporte mensual en Excel con todos los tickets,
    lo envía por mail y borra los tickets de la BD.
    """
    if not MAIL_USERNAME or not MAIL_PASSWORD or 'TU_EMAIL' in MAIL_USERNAME:
        print("ERROR: La configuración de email (MAIL_USERNAME, MAIL_PASSWORD) no está definida.", file=sys.stderr)
        return {
            'success': False,
            'error': 'El servidor no tiene la configuración de email (MAIL_USERNAME, MAIL_PASSWORD) definida.'
        }
    
    conn = get_db()
    c = conn.cursor()
    c.execute('SELECT * FROM tickets ORDER BY date_uploaded DESC')
    tickets = c.fetchall()
    
    if not tickets:
        conn.close()
        return {
            'success': False,
            'error': 'No hay tickets para exportar.'
        }
    
    # --- 1. Crear el Excel con los datos CORRECTOS ---
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Tickets'
    
    # Encabezados 
    headers = ['ID Ticket', 'Usuario', 'Negocio', 'Total Compra', 'Fecha Carga',
           'Producto', 'Cantidad', 'Precio Unitario', 'Precio Total']
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    row = 2
    for ticket in tickets:
        try:
            items = json.loads(ticket['items_json']) if ticket['items_json'] else []
        except:
            items = []

        if not items:
            ws.cell(row=row, column=1, value=ticket['id'])
            ws.cell(row=row, column=2, value=ticket['nombre_usuario'])
            ws.cell(row=row, column=3, value=ticket['nombre_negocio'])
            ws.cell(row=row, column=4, value=ticket['precio_total_compra'])
            ws.cell(row=row, column=5, value=ticket['date_uploaded'])
            row += 1
            continue

        for i, it in enumerate(items):
            ws.cell(row=row, column=1, value=ticket['id'] if i == 0 else None)
            ws.cell(row=row, column=2, value=ticket['nombre_usuario'] if i == 0 else None)
            ws.cell(row=row, column=3, value=ticket['nombre_negocio'] if i == 0 else None)
            ws.cell(row=row, column=4, value=ticket['precio_total_compra'] if i == 0 else None)
            ws.cell(row=row, column=5, value=ticket['date_uploaded'] if i == 0 else None)
            ws.cell(row=row, column=6, value=it.get('producto'))
            ws.cell(row=row, column=7, value=it.get('cantidad'))
            ws.cell(row=row, column=8, value=it.get('precio_unitario'))
            ws.cell(row=row, column=9, value=it.get('precio_total'))
            row += 1
    
    # Ajustar ancho de columnas
    for col in ws.columns:
        max_length = 0
        column_letter = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Guardar temporalmente el archivo
    filename = f'tickets_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    filepath = os.path.join(UPLOAD_FOLDER, filename)
    
    server = None
    
    try:
        wb.save(filepath)
        
        # --- 2. Enviar por mail ---
        msg = MIMEMultipart()
        msg['From'] = MAIL_USERNAME
        msg['To'] = email
        msg['Subject'] = f'Reporte de Tickets - {datetime.now().strftime("%B %Y")}'
        
        body = 'REPORTE DE TICKETS DEL MES ADJUNTO'
        msg.attach(MIMEText(body, 'plain'))
        
        with open(filepath, 'rb') as attachment:
            part = MIMEBase('application', 'octet-stream')
            part.set_payload(attachment.read())
            encoders.encode_base64(part)
            part.add_header('Content-Disposition', f'attachment; filename= {filename}')
            msg.attach(part)
        
        # Conectar y enviar
        print(f"Conectando a {MAIL_SERVER}:{MAIL_PORT}...")
        server = smtplib.SMTP(MAIL_SERVER, MAIL_PORT)
        server.starttls()
        server.login(MAIL_USERNAME, MAIL_PASSWORD)
        print("Enviando email...")
        server.send_message(msg)
        print("Email enviado exitosamente.")
        server.quit()
        
        # --- 3. Borrar tickets ---
        print("Borrando tickets de la base de datos...")
        c.execute('DELETE FROM tickets')
        conn.commit()
        conn.close()
        print("Tickets borrados.")
        
        return {
            'success': True,
            'message': 'Reporte enviado y tickets borrados exitosamente.'
        }
        
    except Exception as e:
        print(f"Error en export_monthly: {e}", file=sys.stderr)
        conn.close()
        return {
            'success': False,
            'error': f'Error: {str(e)}'
        }
            
    finally:
        # --- 4. Limpieza ---
        if server:
            try:
                server.quit()
            except:
                pass
        
        # Borrar el archivo Excel local sin importar si falló o no
        if os.path.exists(filepath):
            try:
                os.remove(filepath)
                print(f"Archivo temporal {filepath} eliminado.")
            except Exception as e:
                print(f"Error al limpiar el archivo {filepath}: {e}", file=sys.stderr)
